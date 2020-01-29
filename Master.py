# -*- coding: utf-8 -*-
"""
@author: Edson Quintero
email: eds07n@ciencias.unam.mx
"""
from datetime import datetime as Datetime
from math import isnan   
from os import listdir
from os.path import exists, isfile, join
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import operator as op
import datetime 
import calendar


global site_folders, headers_file, variables_file, tasks_files, L1_filters
  
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#

### logger_site_folders
######## Where are raw data folders stored?. 
### site_folders
######## Where are the site folders stored?. 
### headers_file
######## File containing the site headers info.
### variables_file
######## File containing the information of the desired variables.
### tasks_Meteo_LX_file
######## Where is the file containing the task tables for each site? (L0 or L1).
### L1_filters
######## File containing the L1 filters.
#-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-D-E-F-I-N-E-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*   
# Where is the logger data stored?
logger_site_folders = '/home/met/data/'
# Where are the L0 folders going to be created?
site_folders = '/home/met/data/'
# 'Headers.xlsx' file path
headers_file = '/home/met/Documents/XLSX/Headers.xlsx'
# 'Site_variables.xlsx' file path
variables_file = '/home/met/Documents/XLSX/Site_variables.xlsx' 
# Where are the 'Tasks_Meteo_X.xlsx' files stored?
tasks_files = '/home/met/Documents/XLSX/' 
# 'L1_filters.xlsx' file path
L1_filters = '/home/met/Documents/XLSX/L1_filters.xlsx'
#-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*    

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#          
# sigdigits(x)
# Returns the accuracy of a given number (x) 
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

def sigdigits(x):   
    
    if isnan(x):
        return np.nan
    else:
        if '.' in str(x):
            return len( ("%s" % float(x)).split('.')[1])
        else:
            return 0

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#                     
# df_accuracy(df, variable)
# Returns the accuracy of a dataframe's column (variable)
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  
            
def df_accuracy(df, variable):
    
    return int(df[variable].apply(sigdigits).mode()[0]) 
  
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#                     
# circ_mean_np(angles,azimuth=True)
# To calculate the mean (average) for points along a circle using the angular measure to these points from the center of the circle 
### Author: Dan.Patterson@carleton.ca
### https://community.esri.com/blogs/dan_patterson/2016/01/10/circular-mean-for-directional-data

### angles: list
########  Angles in degrees
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  
     
def circ_mean_np(angles,azimuth=True):  

    rads = np.deg2rad(angles)  
    av_sin = np.nanmean(np.sin(rads))  
    av_cos = np.nanmean(np.cos(rads))  
    ang_rad = np.arctan2(av_sin,av_cos)  
    ang_deg = np.rad2deg(ang_rad)  
    if azimuth:  
        ang_deg = np.mod(ang_deg,360.)  
        
    return ang_deg  
    
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#
# youngest_date(site)
# Returns the youngest year & month of the given site

### site: str
########  Station code name.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# min_date: youngest date found in the L0 files 
# route: route of the L0 site files
# onlyfiles: files contained in 'route1'
# s_date: date of each file in 'onlyfiles'
# year: youngest year found
# month: youngest month found

def youngest_date(site, sf = site_folders):

    # auxiliar date
    min_date = Datetime.strptime('9999-01', '%Y-%m')
    # where are the L0 files?
    route = join(sf, site, 'L0') 
    # all files list
    onlyfiles = [f for f in listdir(route) if isfile(join(route, f))]
    # for loop to exclude the 'STAT' files    
    for s in onlyfiles:
            if 'STAT' not in s and 'L0' in s:
                # obtain the date of each file
                s_date = Datetime.strptime(s[0:7], '%Y-%m')
                # compare dates to find the youngest
                if min_date > s_date:
                    min_date = s_date
    # get the youngest year found
    year = min_date.strftime('%Y')
    # get the youngest month found
    month = min_date.strftime('%m')  
    
    return year, month

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# add_header(site, file_path, data = 'Meteorological')
# Adds the station header to a specific file

### site: str
########  Station code name. 
### file_path: str
########  Path of the file to which the header is added.
### area: str, default 'Meteorological'
########  Specifies the name of the table. 'Missing' should be used if the function
########  generates the missing data month report.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# headers: dataframe that contains the header information for each site
# line: header that is going to be used
# N: north coordinate of the site
# name: name of the atmospheric observatory / meteorological station 
# Headers_file: location of the 'Headers.xlsx' file
# state: state where the station is located
# UT: universal time 
# W: west coordinate of the site

def add_header(site, file_path, area = 'Meteorological'):
    
    # read the 'Header.xlsx' file
    headers = pd.read_excel(headers_file, converters = {'Name': str, 'State': str, 'North': str, 'West': str, 'MASL': str, 'UT': str})
    # set the site as the index                
    headers = headers.set_index([('Site')]) 
    # extract the header info for each site
    name =  headers['Name'].ix[site]
    state = headers['State'].ix[site]
    N = headers['North'].ix[site]
    W = headers['West'].ix[site]
    masl = headers['MASL'].ix[site]
    UT = headers['UT'].ix[site]
    # header that is going to be used
    #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-HEADER-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-           
    line = 'Red Universitaria de Observatorios Atmosfericos (RUOA)\r\n'+ name + ' (' + site + '), ' + state + '\r\nLat ' + N + ' N, Lon ' + W +' W, Alt ' + masl + ' masl\r\nTime UTC-' + UT + 'h\r\n' + area +' data\r\n '
    #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*
    # here the files are re-opened to add the missing lines   
    with open(file_path, 'r+', encoding = 'latin-1') as f:
        content = f.read()
        f.seek(0,0)
        f.write(line.rstrip('\r\n') + '\r\n' + content)
    # delete the 9th line (the 9th line is a line full of commas)        
    #lines = []
    #with open(file_path, 'r') as f:
    #    lines = f.readlines()
    #with open(file_path, 'w') as f:
    #    f.writelines(lines[:8] + lines[9:])  

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# missing_data_report(site, incoming_route, skip_rows = [0,1,2,3,4,5,7])
# Generates the missing data month report

### site: str
########  Station code name. 
### file_path: str
########  Path of the file to which you want to create the missing data month report. 
### skip_rows: array, default '[0,1,2,3,4,5,7]'
########  Specifies the rows to be skipped when the 'file_path' file is read.
########  '[0,1,2,3,4,5,7]' must be maintained when files that possesses the RUOA header
########  are read, otherwise it has to be changed.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# count_nan: datafame with the daily missing data
# df: datafame in which the null data will be counted
# outcoming_route: path of the missing data report

def missing_data_report(site, file_path, skip_rows = [0,1,2,3,4,5,7]):
    
    # read the dataframe    
    df = pd.read_csv(file_path, skiprows = skip_rows, na_values=['null'])
    # identify if the 'RECORD' column is in df   
    if [col for col in df.columns if 'RECORD' in col]:
        # If it is, then delete it
        del df['RECORD']
    # specify the date column
    df['TIMESTAMP'] = pd.to_datetime(df['TIMESTAMP'], dayfirst = False)
    # set the 'Timestamp' column as the index   
    df = df.set_index(['TIMESTAMP'])
    # count missing data 
    count_nan = 1440 - df.resample('D').count() 
    # define the route where the missing data report is going to be stored
    #parts = (join(site_folders, site, file_path.split('/')[5], 'stat', file_path.split('/')[6]))
    parts = (join(file_path.split('L0')[0], 'L0', 'stat', file_path.split('/')[6]))
    outcoming_route = parts.split('.')[0] + '_STAT.csv'
    print(outcoming_route)
    #outcoming_route = parts[0] + '_STAT.csv'
    # save the missing data report   
    count_nan.to_csv(outcoming_route, line_terminator = '\r\n', na_rep='NaN')
    # add the header to the file
    add_header(site, outcoming_route, area = 'Missing')

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# fill_missing_measurements(df, Time_column = 'TIMESTAMP')
# Deletes duplicate measurements, sort the data by date and the missing measurements
# are added (NaN rows are placed instead).  

### df: Dataframe
########  Dataframe to be worked. 
### Time_column: str, default 'TIMESTAMP'
######## Name of the column that contains the measurement dates. This argument doesn't
######## need to be specified for meteorological data.
### mtd: {None, ‘backfill’/’bfill’, ‘pad’/’ffill’, ‘nearest’}, optional. See pandas.DataFrame.reindex documentation
######## Method to use for filling holes in reindexed DataFrame.
### tol: str, optional
######## Maximum distance between original and new labels for inexact matches. See pandas.DataFrame.reindex documentation
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# oldest: the oldest date in df
# youngest: the youngest date in df 
# begin: the first measurement of the oldest month
# final: the last measurement of youngest
# dates_range: range of dates from 'begin' to 'final' (minute frequency).  

def fill_missing_measurements(df, Time_column = 'TIMESTAMP', mtd = None, tol = None, freq = '1min'):
    
    # detecting duplicated data
    if np.count_nonzero(df.duplicated([Time_column])) != 0:
        # delete data with duplicate date
        df = df.drop_duplicates([Time_column])
    # sort data by date
    df = df.sort_values([Time_column])
    # reset index
    df = df.reset_index(drop = True)
    # the oldest and the youngest dates in the data
    oldest = df[Time_column][0]    
    youngest = df[Time_column][df.index[-2]] 
    # obtain begin
    begin = oldest.strftime('%Y') + '/' + oldest.strftime('%m') + '/1 00:00:00'
    # obtain final
    ### Optional line:
    final = youngest.strftime('%Y/%m/%d %H')
    # range of dates, every minute 
    dates_range = pd.date_range(start = begin, end = final, freq = freq)
    # indexed appended_data with the dates
    df = df.set_index([Time_column])
    # reindexed the appended data
    # df was filled with NaN where there werent data
    #df = df.reindex(dates_range)
    df = df.reindex(dates_range, method = mtd, tolerance = tol)
    
    return df, oldest, youngest

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# record(site, Rrd, oldest, youngest)
# Checks if the record file exists. If not, the file is created and 'FileDate' is chosen,
# otherwise the file is opend and 'FileDate' is extracted. See 'FileDate' below. 
 
### site: str
########  Station code name. 
### Rrd: str
########  Path where the Record file should be stored. 
### oldest: Timestamp (pandas)
########  Oldest date in the data. It can be obtain from the 'fill_missing_measurements' function.
### youngest: Timestamp (pandas)
########  Youngest date in the data. It can be obtain from the 'fill_missing_measurements' function.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# F: Auxiliary variable that helps to know if the record file exists
# FileDate: Date that would be considered the starting date
# R1: Dataframe that contains the record of the previous FileDates
  
def record(site, Rrd, oldest, youngest):
    
    # if it doesn't exist, then begin with the fist date of the whole data
    if not exists(Rrd):      
        # Does the file exist?
        F = 0;
        # the oldest date allowed
        FileDate = oldest
        # and create the record file
        R1 = {'Last ' + site + ' Updload Date' : [youngest]}
        # convert it to Dataframe
        R1 = pd.DataFrame(R1)
        # save the record file
        R1.to_csv(Rrd, line_terminator = '\r\n', index=False)
    # if it exists, then begin with the last uploaded date
    else:
        # Does the file exist?
        F = 1;
        # read the file
        R1 = pd.read_csv(Rrd)
        # set datetime64 into the 'Last Updload Date' column
        R1['Last '+ site +' Updload Date'] = pd.to_datetime(R1['Last '+ site +' Updload Date'])  
        # find the oldest uploaded data
        FileDate = max(R1['Last '+ site +' Updload Date'])  
    
    return F, R1, FileDate

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# write_record(site, youngest, Rrd, F, R1)
# Report the last date contained in the dataframe
 
### site: str
########  Station code name. 
### youngest: Timestamp (pandas)
########  Youngest date in the data. It can be obtain from the 'fill_missing_measurements' function.
### Rrd: str
########  Path of the Record file. 
### F: int
########  Variable returned from the 'record' function.  
### R1: Dataframe
########  Variable returned from the 'record' function. 
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# R2: dataframe with all the record's dates
  
def write_record(site, youngest, Rrd, F, R1):
    
    # Did the record file exist?
    if F == 1:
        ### Create the new Dataframe
        R2 = {'Last ' + site + ' Updload Date' : [youngest]}
        # convert it to Dataframe
        R2 = pd.DataFrame(R2)
        # append the new date
        R2  = R2.append(R1, ignore_index=True)
        # save the record file
        R2.to_csv(Rrd, line_terminator = '\r\n', index=False)
        
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# col_name(site, df, Timestamp = False, Record = False)
# Returns the columns that are in the Dataframe 'df' and in the file 'Variables_file'
        
### site: str
########  Station code name. 
### df: Dataframe
########  Dataframe to be worked.
### Timestamp: bool
########  Do you want to add the 'Timestamp' column.
### Record: bool
########  Do you want to add the 'Record' column.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# site_variables: 'site_variables.xlsx' dataframe
# V: list of site variables extracted from 'site_variables.xlsx'
# U: list of variables units extracted from 'site_variables.xlsx'
# a: list of valid variables
# variables: final list of site variables
# units: final list of variables units

def col_name(site, df, Timestamp = False, Record = False, sheet_name = 'met'):
    
    # read the 'site_variables.xlsx' file
    site_variables = pd.read_excel(variables_file, header = [0,1], sheetname = sheet_name)
    # extract the variable & unit names
    V, U = zip(*site_variables.columns.tolist())
    # identify the columns marked with an 'X'
    a = pd.notnull(site_variables.ix[site]).tolist()
    # lists where the column names (units) are going to be stored
    variables = []; units = []
    # loop to create the list
    for i in range(0,len(site_variables.columns)):
        # if the column name (unit) is in the 'site_variables.xlsx' file & in the dataframe
        # append that column name (unit)
        if a[i] == True and V[i] in df.columns: # and not pd.isnull(df[V[i]]).all():
            variables.append(V[i])
            units.append(U[i])
    # change the encoding format to string format
    #variables = [x.encode('UTF8') for x in variables]
    #units = [x.encode('UTF8') for x in units]
    # if Record == True, then add the 'RECORD' column
    if Record:
        variables = ['RECORD'] + variables
        units = ['RN'] + units
    # if Timestamp == True, then add the 'TIMESTAMP' column
    if Timestamp:
        variables = ['TIMESTAMP'] + variables
        units = ['yyyy-mm-dd HH:MM:SS'] + units
    # return the two lists
    
    return variables, units
  
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# merge_files(folder_path, include = [''], exclude = ':', skip_rows = [])
# Joins different dataframes in one
            
### folder_path: str
########  Path of the folder where the files to be merged are stored. 
### include: list, default ['']
########  If you only want to join files that contain words in common, 'include' must be a list with those words.
########  WARNING : If you only want to join files that contain just one word in common, 'include' MUST be a list with one element 
########  By default all the files contained in the folder are going to be merged.            
### exclude: str, default ':'
########  If you only want to join files that NOT contain a word in common, 'exclude' must be that word.
########  By default all the files contained in the folder are going to be merged.            
### skip_rows: array, default '[]'
########  Specifies the rows to be skipped when the files to be read are read.
########  By default no rows are skipped.              
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# appended_data: resulting dataframe from merging desired the files 
# route: route of each file to be merged
# onlyfiles: list of files within the 'folder_path'

def merge_files(folder_path, include = [''], exclude = ':', skip_rows = []):
    
    # onlyfiles is a list of the files of each site folder
    onlyfiles = [f for f in listdir(folder_path) if isfile(join(folder_path, f))]
    # appended_data is an empty dataframe that is going to be filled with the
    # complete minute site data
    appended_data = []
    # the loop for reading each 'minuto' files:
    for s in onlyfiles:
        if any(i.lower() in s.lower() for i in include) and exclude.lower() not in s.lower():
            route = join(folder_path,s)
            print (route) 
            data = pd.read_csv(route,sep=',',encoding='latin-1', skiprows = skip_rows, na_values = ['NAN','null','nan',' '])
            # appending the data
            appended_data.append(data)
    # concat the dataframes
    appended_data = pd.concat(appended_data) 
    
    return appended_data
  
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# save_month(folder_path, site, df, version = 'L0', extension = '.csv', na_representation = 'null', Timestamp = False, Record = False, MDR = 'True')
# Groups by month the dataframe and saves the month data
            
### folder_path: str
########  Path of the folder where the files are going to be stored. 
### df: Dataframe
########  Dataframe to be worked.     
### version: str
########  {'L0', 'L1'}, it's just used to name the file.
### extension: str
########  {'.dat', '.csv', '.txt'}, idesire extention of the created files.     
### na_representation: str
########  {'null', 'NaN'}, missing data representation.     
### Timestamp: bool
######## Used in the 'col_name' function, see 'col_name'.  
### Record: bool
######## Used in the 'col_name' function, see 'col_name'.  
### MDR: bool
######## Do you want to create missing data reports?.                          
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# grouped_year_month: Object that contains the month groups
# Gdf: Dataframe of each month
# date: (year (Y), month (M)) date of each group (Gdf)
# filename: name of each month file

def save_month(folder_path, site, df, version = 'L0', extension = '.csv', na_representation = 'null', Timestamp = False, Record = False, MDR = True, sheet_name = 'met', area = 'Meteorological'):
    
    # group the 'df' by year & month
    grouped_year_month = df.groupby([lambda x: x.year, lambda x: x.month]) 
    ### save the groups 
    for group in grouped_year_month:
        # extract the date & the data of each group
        date, Gdf = group; 
        # extract the year & the month of the date        
        Y, M = date
        # determinate the empty columns in the dataframe
        cols, units = col_name(site, Gdf, Timestamp, Record, sheet_name)
        Gdf = Gdf[cols]
        # change the header
        header = [['TIMESTAMP'] + cols, ['yyyy-mm-dd HH:MM:SS'] + units]
        ### name the file     
        filename = str(datetime.date(Y, M,1))[:7] + '-' + site + '_' + version + extension
        # reset the dataframe index
        Gdf  = Gdf.reset_index()
        ### insert the header ###
        # minute table    
        Gdf.columns  = pd.MultiIndex.from_tuples(list(zip(*header)))    
        # save the month files
        round(Gdf, 3).to_csv(join(folder_path, filename), na_rep = na_representation,  index=False, encoding = 'latin-1')   
        ### header
        add_header(site,join(folder_path, filename), area = area) #####AREA
        ### missing data report 
        if MDR:
            missing_data_report(site, join(folder_path, filename))
 
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# columns_correction(df)
# Identifies and fixes the problem with the rain & visibility columns (Meteorological Data, L1)
 
### df: Dataframe
########  Dataframe to be worked.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# Visibility_name: Visibility column correct name
  
def columns_correction(df):

    # RAIN
    # change 'Rain_Avg' for 'Rain_Tot'
    if [col for col in df.columns if 'Rain_Avg' in col]:
        del df['Rain_Avg'] 
    # VISIBILITY
    # change 'Visibility' for 'Visibility_Avg' if needed
    if min(df.index) >= pd.to_datetime('2017-03-01'):
        if [col for col in df.columns if 'Visibility' in col]:
            del df['Visibility_Avg'] 
            df.rename(columns={'Visibility':'Visibility_Avg'}, inplace=True) 
    else:
        if [col for col in df.columns if 'Visibility_Avg' in col]:
            del df['Visibility'] 
            df.rename(columns={'Visibility_Avg':'Visibility'}, inplace=True)
    # identify the correct visibility column
    if 'Visibility_Avg' in (df.columns):
        Visibility_name = 'Visibility_Avg'
    elif 'Visibility' in (df.columns):
        Visibility_name = 'Visibility'
    else:
        Visibility_name = False
  
    return (df, Visibility_name)

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# L1_extra_filters(site, df)
# Applies extra L1 filters to a Dataframe. Returns the 'clean' Dataframe.
 
### site: str
########  Station code name. 
### df: Dataframe
########  Dataframe to be worked.  
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# site_filters: Dataframe with all the extra filters
# var1 & var2: variable to be deleted if 'var2' meets the given condition (operators, values)
# var2: Upper limit of the site
# operators: Dates that don't satisfy the lower range
# values: Dates that don't satisfy the upper range
  
def L1_extra_filters(site, df):
    
    # dictionary that contains all the operators 
    operator_dict = {'>': op.gt, '>=': op.ge, '<': op.lt, '<=': op.le, '=': op.eq}
    # read the 'L1_filters' file ('Extra Filters' sheet)
    site_filters = pd.read_excel(L1_filters, sheetname = 'Extra Filters', index_col = 0)
    # identify the 'var1', 'var2', 'operators' and 'values'
    var1 = site_filters['Delete'].ix[site]
    var2 = site_filters['If Variable'].ix[site]
    operators = site_filters['Operator'].ix[site]
    values = site_filters['Value'].ix[site]
    # loop to apply each extra filter
    for col2del, col, operator, val in zip(var1, var2, operators, values):
        # identify the index of the data that meets the given condition
        index = df[operator_dict[operator](df[col],val)].index
        # eliminate the data indicated by the 'index'
        df.set_value(index, col2del, np.nan )
    
    return df
        
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# L1_filters(site, df)
# Applies L1 filters to a Dataframe. Returns the 'clean' Dataframe.
 
### site: str
########  Station code name. 
### df: Dataframe
########  Dataframe to be worked.  
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# site_filters: Dataframe with all the filters
# Lower_Lim: Lower limit of the site 
# Upper_Lim: Upper limit of the site
# Lower_index: Dates that don't satisfy the lower range
# Upper_index: Dates that don't satisfy the upper range
  
def L1_filter(site, df, extra_filters = False, sheet_name = 'met'):
    
    # do you want to apply extra filters? (I)
    if extra_filters:
       df = L1_extra_filters(site, df)   
    # read the 'L1_filters' file
    site_filters = pd.read_excel(L1_filters, header = [0,1], sheetname = 'Filters_' + sheet_name)
    # loop to define each site filters
    for col in df.columns:
        # identify 'Lower_Lim' & 'Upper_Lim'
        Lower_Lim = site_filters[col]['Lower_Lim'].ix[site]
        Upper_Lim = site_filters[col]['Upper_Lim'].ix[site]
        # delete data outside the limints        
        # lower limits
        Lower_index = df[df[col] < Lower_Lim].index
        df.set_value(Lower_index, col, np.nan )
        # upper limits
        Upper_index = df[df[col] > Upper_Lim].index        
        df.set_value(Upper_index, col, np.nan )
    # do you want to apply extra filters? (II)
    if extra_filters:
       df = L1_extra_filters(site, df)
        
    return df
          
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# df_mean(df, Visibility_name, rule)
# Gets averages of a dataframe (df). 
  
### df: Dataframe
########  Dataframe to be worked.   
### rule : str
######## The offset string or object representing target conversion. E.g., '15Min' or '5Min', default set in '60Min'.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# 
  
def df_mean(df, site, rule = '60Min', sheet_name = 'met'):
    
    area_dict = {'met': '', 'cael': '_cael'}
    # IMPORTANT NOTE: 'dict' is the dictionary that contains the instruction of how 
    # each column has to be resampled (i.e., which function should be used to resampled). 
    # If a column that is not in 'dict' needs to be resampled, it has to be added into 
    # the dictionary like this: 'column_name': 'function'
    #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*   
    dict = {'Temp_Avg': np.nanmean, 'RH_Avg': np.nanmean, 'WSpeed_Avg': np.nanmean, 'WSpeed_Max': np.max, 'WDir_Avg': circ_mean_np, 'Rain_Tot': np.nansum, 'Press_Avg': np.nanmean, 'Rad_Avg': np.nanmean, 'Visibility': np.nanmean, 'Visibility_Avg': np.nanmean, 'E_Avg': np.nanstd}
    #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*   
    # get the first hour before the first measurement of the dataframe (date1)
    date1 = (min(df.index) - relativedelta(hours = 1) + relativedelta(minutes = 1)).strftime('%Y-%m-%d %H:%M')
    # get the last measurement of the dataframe
    date2 = max(df.index)
    # try to read the 'date1' month file
    try:
        route1 = join('/home', sheet_name, 'data', site, 'L1', 'Minuto', date1[0:7] + '-' + site + '_minuto_L1' + area_dict[sheet_name] + '.csv') 
        df1 = pd.read_csv(route1, sep=',',skiprows=[0,1,2,3,4,5,7], encoding='latin-1', na_values=['null','NAN','nan'])     
        # specify the date column & set the 'Timestamp' column as the index   
        df1['TIMESTAMP'] = pd.to_datetime(df1['TIMESTAMP'], dayfirst = False)
        df1 = df1.set_index(['TIMESTAMP'])
        # if the file exists, concat the last hour of the file with the df
        df = pd.concat([df1[df1.index >= date1], df])
    except:
        pass
    # find out which is the last hour of the df's month (last_hr_month)
    last_hr_month = pd.to_datetime(date2.strftime('%Y-%m') + '-' + str(calendar.monthrange(date2.year, date2.month)[1]) + ' 23:00')
    # is 'date2' older than 'last_hr_month'?
    df = df[df.index <= last_hr_month]
    # A: Yes
#    if date2 > last_hr_month:    
#        # set 'date2' as the last hour of the next month
#        next_month = last_hr_month + relativedelta(months=1)
#        date2 = next_month.strftime('%Y-%m') + '-' + str(calendar.monthrange(next_month.year, next_month.month)[1]) + ' 23:00'
#            # try to read the 'date2' month file
#        try:
#            route2 = join(site_folders, site, 'L1', 'Minuto', date2[0:7] + '-' + site + '_minuto_L1.csv')
#            df2 = pd.read_csv(route2, sep=',',skiprows=[0,1,2,3,4,5,7], encoding='latin1', na_values=['null','NAN','nan']) 
#            # specify the date column & set the 'Timestamp' column as the index   
#            df2['TIMESTAMP'] = pd.to_datetime(df2['TIMESTAMP'], dayfirst = False)
#            df2 = df2.set_index(['TIMESTAMP']) 
#            # if the file exists, concat the last hour of the file with the df
#            df = pd.concat([df, df2[df2.index <= date2]])
#        except:
#            pass        
    # identify if the 'WDir_SD' column is in the df   
    if [col for col in df.columns if 'WDir_SD' in col]:
        # If it is, then delete it
        del df['WDir_SD'] 
    # obtain the columns order
    order = df.columns   
    # read the 'L1_filters' file
    NMeasurements = pd.read_excel(L1_filters, header = [0], sheetname = 'NMeasurements_' + sheet_name)
    # extract the correct columns from 'dict'
    dict1 = {x:dict[x] for x in order}
    # resample the df, (mean frecuency given by 'rule')
    mean_df = df.resample(rule = rule, closed = 'right', label = 'right').agg(dict1)
    # order the columns
    mean_df = mean_df[order]
    # count the missing measurements
    C = df.resample(rule = rule, closed = 'right', label = 'right').count()
    for col in mean_df.columns:
        # find the columns accuracy 
        try:
            round_n = df_accuracy(df, col)
            mean_df[col] = mean_df[col].round(round_n)
        except:
            pass
        # delete invalid means
        idx = C[C[col] < NMeasurements[col][0]].index
        mean_df.set_value(idx,col, np.NaN )
        
    return mean_df

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#
# tasks_values_L0(Task_List, i)
# Extracts the task arguments from 'Tasks_Met_L0_file'

### Task_List: Dataframe
######## Dataframe read from 'Tasks_Met_L0_file'.
### i: int
######## Index of each task.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#

# task: task to realize
# variable: involved variable
# begin: beginning of the period to be corrected
# end: end of the period to be corrected
# old_M: old multiplier
# new_M: new multiplier
# old_b: old intersect
# new_b: new intersect

def tasks_values_L0(Task_List, i):
        
        # Extract all the info.
        task = Task_List['Task'].ix[i]
        variable = Task_List['Variable'].ix[i]
        begin = Task_List['Begin'].ix[i]
        end = Task_List['End'].ix[i]
        old_M = Task_List['Old Multiplier'].ix[i]
        new_M = Task_List['New Multiplier'].ix[i]
        old_b = Task_List['Old Intersect'].ix[i]
        new_b = Task_List['New Intersect'].ix[i]
        
        return task, variable, begin, end, old_M, new_M, old_b, new_b

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#
# tasks_values_L1(Task_List, i)
# Extracts the task arguments from 'Tasks_Met_L1_file'

### Task_List: Dataframe
######## Dataframe read from 'Tasks_Met_LX_file'.
### i: int
######## Index of each task.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#

# task: task to realize
# variable: involved variable
# begin: beginning of the period to be corrected
# end: end of the period to be corrected
# operator: operator
# value: value

def tasks_values_L1(Task_List, i):
        
        # Extract all the info.
        task = Task_List['Task'].ix[i]
        variable = Task_List['Variable'].ix[i]
        begin = Task_List['Begin'].ix[i]
        end = Task_List['End'].ix[i]
        operator = Task_List['Operator'].ix[i]
        value = Task_List['Value'].ix[i]
        
        return task, variable, begin, end, operator, value
     
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# hour2minute(site, variable, begin, end)
# Adds hourly data to the minute table
            
### site: str
########  Station code name. 
### variable: str
########  Name of the column to be worked.  
### begin: Timestamp (pandas)
########  First day to be worked.
### end: Timestamp (pandas)
########  Last day to be worked.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# route0: Where are the hour data files?
# route1: Where are the minute data files?
# Hdf: Dataframe with the hour data
# dates_array: array with the dates to be worked
# df: Dataframe with the minute data

def hour2minute(site, variable, begin, end): 
    
    #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*    
    route0 = join(logger_site_folders, site, 'raw') 
    route1 = join(site_folders, site, 'L0') 
    #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*
    # merge the hour files    
    Hdf = merge_files(route0, skip_rows = [0,2,3], include = ['hora'])
    # extract the desire column    
    Hdf = Hdf[['TIMESTAMP', variable]]
    # convert the TIMESTAMP column into datetime64
    Hdf['TIMESTAMP'] = pd.to_datetime(Hdf['TIMESTAMP'], dayfirst = False)
    # clean the hour data
    Hdf, oldest, youngest = fill_missing_measurements(Hdf, mtd = 'ffill', tol = '1 hr' )   
    # list with dates range
    dates_array = pd.date_range(start = begin, end = end + relativedelta(months = 1), freq = 'M').strftime('%Y-%m')  
    # merge the files indicated in 'dates_array'    
    df = merge_files(route1, skip_rows = [0,1,2,3,4,5,7], include = dates_array, exclude = 'STAT')
    # add the column if it doen't exist    
    if variable not in df: 
        df[variable] = np.nan
    # convert the TIMESTAMP column into datetime64
    df['TIMESTAMP'] = pd.to_datetime(df['TIMESTAMP'], dayfirst = False)    
    # change the df index    
    df = df.set_index(['TIMESTAMP'])
    # insert the hour data into the minute data
    df[variable][begin.strftime('%Y-%m-%d %H'):end.strftime('%Y-%m-%d %H')] = Hdf[variable][begin.strftime('%Y-%m-%d %H'):end.strftime('%Y-%m-%d %H')].values
    # save the new minute data    
    save_month(route1, site, df, extension = '.dat', Record = True)

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# multipliers(site, variable, begin, end, old_M, new_M, old_b, new_b)
# Changes the multipliers for a determinated variable in a given period of time
            
### site: str
########  Station code name. 
### variable: str
########  Name of the column to be worked.  
### begin: Timestamp (pandas)
########  First day to be worked.
### end: Timestamp (pandas)
########  Last day to be worked.
### old_M: float
########  Old multiplier.
### new_M: float
########  New multiplier.
### old_b: float
########  Old intersect.
### new_b: float
########  New intersect.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# route: Where are the L0 files?
# dates_array: array with the months to be worked
# df: dataframe that contains the 'dates_array' data
# mask: dates to be worked
# round_n: accuracy of the variable column

def multipliers(site, variable, begin, end, old_M, new_M, old_b, new_b): 
    
    #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*    
    route = join(site_folders, site, 'L0') 
    #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*
    # list with dates range
    dates_array = pd.date_range(start = begin, end = end + relativedelta(months = 1), freq = 'M').strftime('%Y-%m')  
    # merge the files indicated in 'dates_array'    
    df = merge_files(route, skip_rows = [0,1,2,3,4,5,7], include = dates_array, exclude = 'STAT')
    # convert the TIMESTAMP column into datetime64
    df['TIMESTAMP'] = pd.to_datetime(df['TIMESTAMP'], dayfirst = False)    
    # change the df index    
    df = df.set_index(['TIMESTAMP'])
     # mask1: the first period
    mask = ((df.index >= str(begin)) & (df.index <= str(end)))
    # the corrections for each period 
    df.loc[mask, variable] = (df.loc[mask, variable] - old_b) * (new_M/old_M) + new_b                
    # df.loc[mask, variable] = df.loc[mask, variable] * (new_value/old_value)
    # find the column accuracy    
    round_n = df_accuracy(df, variable) 
    # round the variable column
    df[variable] = df[variable].round(round_n)
    # save the new data    
    save_month(route, site, df, extension = '.dat', Record = True)

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# reworkdates(site, task, begin, counter, version = 'L0') 
#            
### site: str
########  Station code name. 
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

def reworkdates(site, task, begin, counter, version = 'L0'): 
    
    tfile = join(tasks_files, 'Tasks_Meteo_L0.xlsx')
    if counter == 1:
        #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*    
        route1 = join(site_folders, site,'L0', site+'_record.txt')
        #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*
        # read the file
        R = pd.read_csv(route1)
        # set datetime64 into the 'Last Updload Date' column
        R['Last '+ site +' Updload Date'] = pd.to_datetime(R['Last '+ site +' Updload Date'])  
        # delete the record data    
        R = R[0:0]
        # write the desire date
        write_record(site, begin, route1, 1, R)   
        success = False
    else:
        # read the 'Tasks_List.xlsx' file
        Task_List = pd.read_excel(tfile, sheetname = site)
        # convert the 'Beging' and 'End' columns into datetime64 format
        Task_List['Begin'] = pd.to_datetime(Task_List['Begin'], dayfirst = False)
        Task_List['End'] = pd.to_datetime(Task_List['End'], dayfirst = False)
        #    
        idx = Task_List[Task_List['Task'] == task].index[0] 
        # load the excel workbook
        wb = load_workbook(tfile)
        # select the site sheet
        sheet = wb.get_sheet_by_name(site)
        for i in range(1,8):
            sheet.cell(row=(idx+2), column=i).value = ''     
        wb.save(tfile)
        # find the tasks that have to be made again 
        index = Task_List[pd.notnull(Task_List['Corrected Task'])].index.tolist() 
        # indicate if there arent new tasks    
        if not index:
            pass
       # loop for each task    
        else:
            for i in index:  
                task, variable, BEGIN, end, old_M, new_M, old_b, new_b = tasks_values_L0(Task_List, i)
                begin = Datetime.strptime(begin.strftime('%Y-%m'), '%Y-%m')                
                if begin <= end and begin  <= BEGIN:                    
                    tasks_L0(site, task, variable, BEGIN, end, old_M, new_M, old_b, new_b, counter)
                elif begin <= end and begin  > BEGIN:                   
                    tasks_L0(site, task, variable, begin, end, old_M, new_M, old_b, new_b, counter)
        success = True
    counter += 1
    
    return (success, counter)         

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# remove_if(site, variable, begin, end, operator, value)
# Removes values if they meet a condition (operator, value)
            
### site: str
########  Station code name. 
### variable: str
########  Name of the column to be worked.  
### begin: Timestamp (pandas)
########  First day to be worked.
### end: Timestamp (pandas)
########  Last day to be worked.
### operator: str
########  Function corresponding to the intrinsic operators of Python.
### value: str
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# Very similar to the 'multipliers' function (check above).

def remove_if(site, variable, begin, end, operator, value): 
    
    # dictionary that contains all the operators 
    operator_dict = {'>': op.gt, '>=': op.ge, '<': op.lt, '<=': op.le, '=': op.eq}
    #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*    
    route = join(site_folders, site, 'L1') 
    #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*
    # list with dates range
    dates_array = pd.date_range(start = begin, end = end + relativedelta(months = 1), freq = 'M').strftime('%Y-%m')  
    # merge the files indicated in 'dates_array'    
    df = merge_files(join(route, 'Minuto'), skip_rows = [0,1,2,3,4,5,7], include = dates_array, exclude = 'STAT')
    # convert the TIMESTAMP column into datetime64
    df['TIMESTAMP'] = pd.to_datetime(df['TIMESTAMP'], dayfirst = False)    
    # change the df index    
    df = df.set_index(['TIMESTAMP'])
    # mask:
    if operator.lower() == 'all':
        mask = ((df.index >= str(begin)) & (df.index <= str(end)))
    else:        
        mask = ((df.index >= str(begin)) & (df.index <= str(end)) & (operator_dict[operator](df[variable],float(value))))
    # eliminate the desire data                   
    df.loc[mask, variable] = df.loc[mask, variable] * np.NaN
    # calculate the hourly means
    min60 = df_mean(df, site)
    # save the new dfs    
    save_month(join(route, 'Minuto'), site, df, version = 'minuto_L1', MDR = False)
    save_month(join(route, 'Hora'), site, min60, version = 'hora_L1', MDR = False)
    
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# hour2hour(site, variable, begin, end)
# Adds hourly data to the hour table
            
### site: str
########  Station code name. 
### variable: str
########  Name of the column to be worked.  
### begin: Timestamp (pandas)
########  First day to be worked.
### end: Timestamp (pandas)
########  Last day to be worked.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

# Very similar to the 'hour2minute' function (check above).

def hour2hour(site, variable, begin, end): 
    
    #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*    
    route0 = join(logger_site_folders, site, 'raw') 
    route1 = join(site_folders, site, 'L1', 'Hora') 
    #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*
    # merge the hour files    
    Hdf = merge_files(route0, skip_rows = [0,2,3], include = ['hora'])
    # extract the desire column    
    Hdf = Hdf[['TIMESTAMP', variable]]
    # convert the TIMESTAMP column into datetime64
    Hdf['TIMESTAMP'] = pd.to_datetime(Hdf['TIMESTAMP'], dayfirst = False)
    # clean the hour data
    Hdf, _, _ = fill_missing_measurements(Hdf, freq = '60min')   
    # list with dates range
    dates_array = pd.date_range(start = begin, end = end + relativedelta(months = 1), freq = 'M').strftime('%Y-%m')  
    # merge the files indicated in 'dates_array'    
    df = merge_files(route1, skip_rows = [0,1,2,3,4,5,7], include = dates_array, exclude = 'STAT')
    # add the column if it doen't exist    
    if variable not in df: 
        df[variable] = np.nan
    # convert the TIMESTAMP column into datetime64
    df['TIMESTAMP'] = pd.to_datetime(df['TIMESTAMP'], dayfirst = False)    
    # change the df index    
    df = df.set_index(['TIMESTAMP'])
    # insert the hour data into the minute data
    df[variable][begin.strftime('%Y-%m-%d %H'):end.strftime('%Y-%m-%d %H')] = Hdf[variable][begin.strftime('%Y-%m-%d %H'):end.strftime('%Y-%m-%d %H')].values
    # save the new minute data    
    save_month(route1, site, df, version = 'hora_L1', MDR = False)
    
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#
# tasks_L0(site, task, variable, begin, end, old_M, new_M, old_b, new_b, counter)
# Function that excecutes the possible tasks (L0). 
            
### Arguments
########  Specified in each function.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

def tasks_L0(site, task, variable, begin, end, old_M, new_M, old_b, new_b, counter):
    
    if ('Rework' not in task or counter == 1): 
        # print the task to be realized
        print ('*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#')
        print ('Task: ' + str(task))
        print ('Variable: ' + str(variable))
        print ('Begin: ' + str(begin))
        print ('End: ' + str(end))
        print ('Old Multiplier: ' + str(old_M))
        print ('New Multiplier: ' + str(new_M))
        print ('Old Intersect: ' + str(old_b))
        print ('New Intersect: ' + str(new_b))
    ## 'Add Hour Column':  
    if task == 'Add Hour Column':
        hour2minute(site, variable, begin, end)  
        success = True      
        return success      
    ## 'Multipliers':  
    elif task == 'Multipliers':
        multipliers(site, variable, begin, end, old_M, new_M, old_b, new_b)
        success = True      
        return success      
    ## 'Rework Dates':              
    elif task == 'Rework Dates': 
        success = reworkdates(site, task, begin, counter)      
        return success
    ## 'Rework All Dates':              
    elif task == 'Rework All Dates': 
        year, month = youngest_date(site)
        begin = Datetime.strptime(year+'-'+month, '%Y-%m')
        success = reworkdates(site, task, begin, counter)      
        return success

#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#
# tasks_L1(site, task, variable, begin, end, operator, value))
# Function that excecutes the possible tasks (L1). 
            
### Arguments
########  Specified in each function.
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#  

def tasks_L1(site, task, variable, begin, end, operator, value):
     
    # print the task to be realized
    print ('*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#')
    print ('Task: ' + str(task))
    print ('Variable: ' + str(variable))
    print ('Begin: ' + str(begin))
    print ('End: ' + str(end))
    print ('Remove If: ' + str(operator) + ' ' + str(value))
    ## 'Remove If':  
    if task == 'Remove If':
        remove_if(site, variable, begin, end, operator, value) 
        success = True      
        return success
    ## 'Add Hour Column':  
    elif task == 'Add Hour Column':
        hour2hour(site, variable, begin, end)  
        success = True      
        return success  
                        
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# master_of_none(site, counter, tfile) 
# Identifies the tasks that are necessary to perform. 

### site: str
########  Station code name.                    
### counter
########  Auxiliary variable. It's used to know if it is necessary to repeat the loop.
### tfile:
########     
#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#    
# tfile: location of the 'Tasks_Met_LX_file.xlsx' file
# Task_List: dataframe that contains the tasks to be realized
# index: indexes of the tasks that haven't been made
# task: name of the task
# variable: variable to be corrected
# begin: beginning of the period to be corrected 
# end: end of the period to be corrected
# wb: 
# sheet:
    
def master_of_none(site, counter, area = 'Meteo', version = 'L0'):  
    
    # Identify the file to be opened
    tfile = join(tasks_files, 'Tasks_' + area + '_' + version + '.xlsx')
    # read the 'Tasks_List.xlsx' file
    Task_List = pd.read_excel(tfile, sheetname = site)
    # convert the 'Beging' and 'End' columns into datetime64 format
    Task_List['Begin'] = pd.to_datetime(Task_List['Begin'], dayfirst = False)
    Task_List['End'] = pd.to_datetime(Task_List['End'], dayfirst = False) 
    # find the tasks that haven't been made 
    index = Task_List[pd.isnull(Task_List['Corrected Task'])].index.tolist() 
    # indicate if there arent new tasks    
    if not index:
        print('There aren\'t tasks to be realized')  
        success = True
        # pass to the next site (success == True)
        return success, None       
   # loop for each task    
    else:
        for i in index:  
            # task (task) to be carried out from (begining) to (end) 
            # for the variable (variable)
            if version == 'L0':
                column = 9
                task, variable, begin, end, old_M, new_M, old_b, new_b = tasks_values_L0(Task_List, i)
                success = tasks_L0(site, task, variable, begin, end, old_M, new_M, old_b, new_b, counter)
            elif version == 'L1':
                column = 7
                task, variable, begin, end, operator, value = tasks_values_L1(Task_List, i)   
                success = tasks_L1(site, task, variable, begin, end, operator, value)
            # 'Rework' not in task, mark the 'X'
            if 'Rework' not in task:
                # load the excel workbook
                wb = load_workbook(tfile)
                # select the site sheet
                sheet = wb.get_sheet_by_name(site)
                # mark the task that has already been performed
                sheet.cell(row=(i+2), column=column).value = 'X'    
                # save the workbook
                wb.save(tfile)
            # else,  
            else:
                s, c = success
                # repeat the site loop (success == False)
                return s, c
        # pass to the next site (success == True)  
        return success, None
    
def rainfall_check():
    sites = ['altz','cham','erno','jqro','ltux','unam']
    B = []
    for site in sites:
        b = exists(join('/home/met/data/' ,site,'L1',site+'_rain_correction.txt'))
        B.append(b)
    if not np.all(B):
        folder_path = '/home/met/Documents/Datos_1_min_RUOA_lluvia'
        onlyfiles = [f for f in listdir(folder_path) if isfile(join(folder_path, f))]
        for f in sorted(onlyfiles):
            print(f)
            site = f[:4].lower()
            route1 = join(folder_path, f)
            df = pd.read_csv(route1, skiprows=[0], na_values=['null','NAN',' '])    
            df['timestamp'] = pd.to_datetime(df['timestamp'], dayfirst = True, format = '%d/%m/%Y %H:%M:%S')
            df = df.set_index([('timestamp')]) 
            grouped = df.groupby([lambda x: x.year, lambda x: x.month])  
            for (Y, M), G in grouped:
                route2 = '/home/met/data/' + site + '/L1/Minuto/' + str(Y) + '-' + str(M).zfill(2) + '-' + site + '_minuto_L1.csv'
                data = pd.read_csv(route2, sep=',',skiprows=[0,1,2,3,4,5,7],encoding='latin-1',na_values=['null','NAN']) 
                data['TIMESTAMP'] = pd.to_datetime(data['TIMESTAMP'], dayfirst = True)
                data = data.set_index([('TIMESTAMP')]) 
                col = data.columns.values
                data = pd.concat([data, G], axis=1)
                del data['Rain_Tot']
                data = data.rename(columns = {'Rain_corregido':'Rain_Tot'})
                data = data[col]
                min60 = df_mean(data, site)        
                save_month('/home/met/data/' + site + '/L1/Minuto/', site, data, version = 'minuto_L1', MDR = False)
                save_month('/home/met/data/' + site + '/L1/Hora/', site, min60, version = 'hora_L1', MDR = False)
        R = {'The rainfall corrections were made on:' : [Datetime.now()]}        
        R = pd.DataFrame(R)
        sites = ['altz','cham','erno','jqro','ltux','unam']
        for site in sites:
            R.to_csv(join('/home/met/data/' ,site,'L1',site+'_rain_correction.txt'),index=False)    
        
def frequency(df):
    n = len(df)
    if n < 2:
        f = 60.0
    else:
        f = np.diff(df.index.values).min() / np.timedelta64(1, 's')
    return f
    
def Avg_frequencies(df):
    
    del df['RECORD']
    df = df.drop_duplicates(['TIMESTAMP'])
    df['TIMESTAMP'] = pd.to_datetime(df['TIMESTAMP'], dayfirst = False)
    df = df.set_index(['TIMESTAMP'])   
    new_df = df.resample(rule = '1Min', closed = 'right', label = 'right').apply(np.nanmean)
    C = df['E_Avg'].resample(rule = '1Min', closed = 'right', label = 'right').count()       
    new_df['Frequency'] = df['E_Avg'].resample(rule = '1Min', closed = 'right', label = 'right').apply(frequency)  
    idx = C[C / (60 * new_df['Frequency']**(-1)) < 0.5].index  
    new_df.set_value(idx, 'E_Avg', np.NaN )
    del new_df['Frequency']
    
    return new_df
